import pandas as pd
import json
from openpyxl import load_workbook


def embeddings_dict(content):
    #Get element embeddings dict
    CLS = content['features'][0]["layers"][0]["values"]
    #The 13 dimensions corresponding to the 450 generation
    yield_demension = [23, 54, 99, 108, 111, 148, 159, 230, 235, 319, 336, 406, 705]
    tensile_demension = [129, 144, 159, 165, 177, 310, 406, 446, 452, 556, 561, 618, 680]
    elongation_demension = [58, 88, 214, 255, 301, 343, 356, 535, 609, 647, 677, 688, 759]

    embedding_yield = [CLS[i-1] for i in yield_demension]
    #embedding_tensile = [CLS[i-1] for i in tensile_demension]
    embedding_elongation = [CLS[i-1] for i in elongation_demension]
    return embedding_yield, embedding_elongation


def get_last_row(file_path, sheet_name):
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]
    return sheet.max_row


def batch_process_and_write(batch_data):
    processed_batch = [embeddings_dict(item) for item in batch_data]
    yield_em = [item[0] for item in processed_batch]
    elongation_em = [item[1] for item in processed_batch]
    yield_em = pd.DataFrame(yield_em)
    elongation_em = pd.DataFrame(elongation_em)
    print(yield_em)
    print(elongation_em)
    with pd.ExcelWriter("/Users/wangping/Desktop/pythonProject/D-electron/yield_GBR_450_zhang.xlsx", engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer1:
        yield_em.to_excel(writer1, index=False, header=False, startrow=writer1.sheets['Sheet1'].max_row)
    with pd.ExcelWriter("/Users/wangping/Desktop/pythonProject/D-electron/elongation_GBR_450_zhang.xlsx", engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer2:
        elongation_em.to_excel(writer2, index=False, header=False, startrow=writer2.sheets['Sheet1'].max_row)


def main(_):

    with pd.ExcelWriter("/Users/wangping/Desktop/pythonProject/D-electron/yield_GBR_450_zhang.xlsx", engine='openpyxl', mode='w') as writer1:
        pd.DataFrame(columns=["X23", "X54","X99","X108","X111","X148","X159","X230","X235","X319","X336","X406","X705"]).to_excel(writer1, index=False)
    with pd.ExcelWriter("/Users/wangping/Desktop/pythonProject/D-electron/elongation_GBR_450_zhang.xlsx", engine='openpyxl',mode='w') as writer2:
        pd.DataFrame(columns=["X58","X88","X214","X255","X301","X343","X356","X535","X609","X647","X677","X688","X759"]).to_excel(writer2, index=False)

    batch_size = 5
    batch_data = []

    with open(file, 'r') as f:
        for line in f:
            data = json.loads(line)
            batch_data.append(data)
            if len(batch_data) >= batch_size:
                batch_process_and_write(batch_data)
                batch_data = []

    if batch_data:
        batch_process_and_write(batch_data)

if __name__ == "__main__":
    #file = "/Users/wangping/Desktop/pythonProject/D-electron/Result/candidates_all_em.json"
    file = "/Users/wangping/Desktop/pythonProject/extract_features/Verified by existing alloys.json"
    main(file)

